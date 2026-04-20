"""
Append additional demo rows to the two workbooks used in the CEO demo:
  - data/demo.xlsx               (Pawnee / Bob's Burgers universe)
  - data/Coastal_Harbor_Export.xlsx   (Annapolis / Chesapeake universe)

Rows are themed per workbook so the upload demo looks natively consistent.
Safe to re-run: the script skips any entityIdentifier / humanIdentifier
already present in the sheet.
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"


def _existing_ids(ws, col: int) -> set[str]:
    return {row[col - 1] for row in ws.iter_rows(min_row=2, values_only=True) if row[col - 1]}


def _append(ws, rows: list[tuple], id_col: int = 1) -> int:
    seen = _existing_ids(ws, id_col)
    added = 0
    for r in rows:
        if r[id_col - 1] in seen:
            continue
        ws.append(list(r))
        added += 1
    return added


# ---------- PAWNEE / BOB'S BURGERS UNIVERSE (demo.xlsx) ------------------

PAWNEE_ENTITIES = [
    # entityIdentifier, name, legalEntityType, doi, description, revenue,
    # isClosed, DBA, fein, ssn, allowAgencyParticipation
    ("ent018", "Tom’s Bistro", "llc", dt.datetime(2013, 9, 1),
     "Upscale bistro opened by Thomas Haverford; known for truffle shoestring fries", 1_800_000, False,
     "Tom’s Bistro", "46-1122334", None, True),
    ("ent019", "Sweetums Candy Factory", "c_Corp", dt.datetime(1951, 4, 12),
     "Pawnee sugary-confections manufacturing plant — Sweetums subsidiary", 23_000_000, False,
     None, "45-9876544", None, True),
    ("ent020", "Pawnee Commons Construction", "llc", dt.datetime(2011, 2, 3),
     "General contractor specializing in parks and municipal build-outs", 4_400_000, False,
     "Pawnee Commons", "38-2211001", None, True),
    ("ent021", "Ann Perkins Wellness Clinic", "sole_Proprietor", dt.datetime(2015, 1, 15),
     "Primary care and public health consulting practice", 420_000, False,
     None, None, "345-67-8910", True),
    ("ent022", "Gryzzl Pawnee Office", "c_Corp", dt.datetime(2013, 6, 1),
     "Regional tech office; consumer data and wifi services", 18_500_000, False,
     "Gryzzl", "27-5544332", None, True),
    ("ent023", "Eagleton Country Club", "limited_Partnership", dt.datetime(1978, 7, 4),
     "Private country club and event venue in neighboring Eagleton", 6_200_000, False,
     None, "12-3344556", None, True),
    ("ent024", "Pawnee Zoo", "not_For_Profit", dt.datetime(1934, 5, 1),
     "Municipal zoo; petting zoo, small aquarium, and event space", 850_000, False,
     None, "11-8899003", None, True),
    ("ent025", "Seymour’s Bay Fish Market", "llc", dt.datetime(2004, 8, 20),
     "Family-run seafood market and wholesale distribution", 1_150_000, False,
     "Fresh Catch Co.", "47-5566778", None, True),
    ("ent026", "Wharf Taffy Co.", "c_Corp", dt.datetime(1986, 11, 14),
     "Saltwater taffy producer operating on Wonder Wharf", 3_800_000, False,
     "Wharf Taffy", "51-9988776", None, True),
    ("ent027", "Belcher Catering Services", "sole_Proprietor", dt.datetime(2019, 3, 10),
     "Event catering spin-off from Bob’s Burgers", 180_000, False,
     "Belcher Catering", None, "123-45-6788", True),
]

PAWNEE_HUMANS = [
    # humanIdentifier, prefix, first, middle, last, preferred, pronoun, dob,
    # education, occupation, industry, yearStarted, isDeceased, gender,
    # maritalStatus, ssn, licenseNumber, licenseState, firstLicensedDate,
    # allowAgencyParticipation, (+5 trailing None for width=25)
    ("hum032", "Mr.", "Teddy", None, "Blaszczak", "Teddy", "hehim",
     dt.datetime(1962, 10, 5), "high_School", "skilled_Trade_Technician",
     "construction", 1982, False, "male", "divorced",
     "456-78-9123", None, None, None, True, None, None, None, None, None),
    ("hum033", "Ms.", "Mona-Lisa", None, "Saperstein", "Mona", "sheher",
     dt.datetime(1988, 12, 14), "bachelors", "sales_Marketing",
     "entertainment", 2010, False, "female", "single",
     "567-89-1234", None, None, None, True, None, None, None, None, None),
    ("hum034", "Dr.", "Richard", None, "Nygard", "Dr. Saperstein", "hehim",
     dt.datetime(1958, 1, 20), "phd", "healthcare_Professional",
     "healthcare", 1986, False, "male", "married",
     "678-91-2345", None, None, None, True, None, None, None, None, None),
    ("hum035", "Mr.", "Craig", None, "Middlebrooks", "Craig", "hehim",
     dt.datetime(1972, 4, 8), "bachelors", "manager_Supervisor",
     "government", 2002, False, "male", "single",
     "789-12-3456", None, None, None, True, None, None, None, None, None),
    ("hum036", "Ms.", "Tynnyfer", None, "McGraw", "Tynnyfer", "sheher",
     dt.datetime(1989, 7, 19), "bachelors", "administrator_Clerical",
     "government", 2014, False, "female", "single",
     "891-23-4567", None, None, None, True, None, None, None, None, None),
    ("hum037", "Mr.", "Kyle", None, "Nichols", "Kyle", "hehim",
     dt.datetime(1980, 2, 28), "high_School", "laborer",
     "manufacturing", 2001, False, "male", "single",
     "912-34-5678", None, None, None, True, None, None, None, None, None),
    ("hum038", "Ms.", "Brandi", None, "Maxxxx", "Brandi Maxxxx", "sheher",
     dt.datetime(1983, 5, 5), "high_School", "arts_Entertainment_Recreation",
     "entertainment", 2003, False, "female", "single",
     "132-45-6789", None, None, None, True, None, None, None, None, None),
    ("hum039", "Mr.", "Orin", None, None, None, "hehim",
     dt.datetime(1975, 10, 31), "masters", "other",
     "arts", 2005, False, "male", "single",
     "143-56-7891", None, None, None, False, None, None, None, None, None),
    ("hum040", "Mrs.", "Tammy", "II", "Swanson", "Tammy 2", "sheher",
     dt.datetime(1968, 6, 15), "masters", "administrator_Clerical",
     "education", 1995, False, "female", "divorced",
     "154-67-8912", None, None, None, True, None, None, None, None, None),
    ("hum041", "Ms.", "Diane", None, "Lewis", "Diane", "sheher",
     dt.datetime(1972, 11, 4), "masters", "educator_Professor",
     "education", 2000, False, "female", "married",
     "165-78-9123", None, None, None, True, None, None, None, None, None),
    ("hum042", "Mr.", "Burt", None, "Macklin", "Burt", "hehim",
     dt.datetime(1979, 8, 22), "bachelors", "legal_Financial_Professional",
     "entertainment", 2004, False, "male", "married",
     "176-89-1234", None, None, None, True, None, None, None, None, None),
    ("hum043", "Ms.", "Evelyn", None, "Stonefield", "Evelyn", "sheher",
     dt.datetime(1965, 3, 18), "bachelors", "executive_Owner",
     "manufacturing", 1990, False, "female", "married",
     "187-91-2345", None, None, None, True, None, None, None, None, None),
    ("hum044", "Mr.", "Andy", None, "Dwyer", "Andy", "hehim",
     dt.datetime(1982, 9, 9), "high_School", "arts_Entertainment_Recreation",
     "entertainment", 2003, False, "male", "married",
     "198-12-3457", None, None, None, True, None, None, None, None, None),
    # Intentional near-duplicate of hum013 "Andrew Dwyer" to exercise dedupe
    ("hum045", "Mrs.", "Jennifer", None, "Barkley-Wyatt", "Jen", "sheher",
     dt.datetime(1974, 1, 27), "masters", "sales_Marketing",
     "politics", 1998, False, "female", "single",
     "219-23-4568", None, None, None, True, None, None, None, None, None),
    ("hum046", "Mr.", "Perd", None, "Hapley", "Perd", "hehim",
     dt.datetime(1964, 11, 30), "bachelors", "information_Media_Communications",
     "media", 1988, False, "male", "married",
     "221-34-5679", None, None, None, True, None, None, None, None, None),
]

# Pawnee contacts — all use Pawnee IN addresses; entity contacts are BUSINESS,
# human contacts are PERSONAL with the canonical 40-column width.
_PAWNEE_CONTACTS_ENTITY = [
    # (ent_id, name_hint, line1, city, state, phone, website, email)
    ("ent018", "Main Location", "400 Eagleton Way", "Pawnee", "in", 3175552018, "https://tomsbistro.example.com", "info@tomsbistro.example.com"),
    ("ent019", "Plant", "1 Sugar Mill Rd", "Pawnee", "in", 3175552019, "https://sweetums.example.com", "factory@sweetums.example.com"),
    ("ent020", "Yard", "250 Commons Ave", "Pawnee", "in", 3175552020, "https://pawneecommons.example.com", "jobs@pawneecommons.example.com"),
    ("ent021", "Clinic", "500 Wellness Way", "Pawnee", "in", 3175552021, "https://annperkinsclinic.example.com", "hello@annperkinsclinic.example.com"),
    ("ent022", "Office", "900 Tech Park Blvd", "Pawnee", "in", 3175552022, "https://gryzzl.example.com", "pawnee@gryzzl.example.com"),
    ("ent023", "Clubhouse", "1 Fairway Drive", "Eagleton", "in", 3175552023, "https://eagletoncc.example.com", "events@eagletoncc.example.com"),
    ("ent024", "Main Gate", "200 Zoo Ln", "Pawnee", "in", 3175552024, "https://pawneezoo.example.com", "info@pawneezoo.example.com"),
    ("ent025", "Market", "45 Dockside Rd", "Seymour’s Bay", "ca", 3105552025, "https://seymoursfish.example.com", "orders@seymoursfish.example.com"),
    ("ent026", "Shop", "2 Wonder Wharf", "Seymour’s Bay", "ca", 3105552026, "https://wharftaffy.example.com", "hello@wharftaffy.example.com"),
    ("ent027", "HQ", "123 Ocean Ave", "Seymour’s Bay", "ca", 3105552027, None, "catering@bobsburgers.example.com"),
]

_PAWNEE_CONTACTS_HUMAN_PERSONAL = [
    # (hum_id, line1, city, state, phone, email)
    ("hum032", "800 Lumberyard Ln", "Pawnee", "in", 3175553032, "teddy@example.com"),
    ("hum033", "77 Party Plaza", "Pawnee", "in", 3175553033, "mona@saperstein.example.com"),
    ("hum034", "11 Clinic Row", "Pawnee", "in", 3175553034, "drsap@example.com"),
    ("hum035", "350 Main Street", "Pawnee", "in", 3175553035, "craig.m@pawneeparks.example.com"),
    ("hum036", "222 Admin Ln", "Pawnee", "in", 3175553036, "tynnyfer@pawnee.example.com"),
    ("hum037", "1 Sugar Mill Rd", "Pawnee", "in", 3175553037, "kyle@sweetums.example.com"),
    ("hum038", "77 Party Plaza", "Pawnee", "in", 3175553038, "brandi@saperstein.example.com"),
    ("hum039", "18 Free Spirit Pl", "Pawnee", "in", 3175553039, "orin@example.com"),
    ("hum040", "900 Eagleton Hts", "Eagleton", "in", 3175553040, "tammy2@eagletonlibrary.example.com"),
    ("hum041", "450 University Row", "Pawnee", "in", 3175553041, "diane.lewis@example.com"),
    ("hum042", "350 Main Street", "Pawnee", "in", 3175553042, "burt.macklin@fbi.example.com"),
    ("hum043", "1 Sugar Mill Rd", "Pawnee", "in", 3175553043, "evelyn@stonefield.example.com"),
    ("hum044", "222 Hideaway Trail", "Pawnee", "in", 3175553044, "andy.dwyer@example.com"),
    ("hum045", "450 Campaign Blvd", "Pawnee", "in", 3175553045, "jen@barkley.example.com"),
    ("hum046", "500 Press Row", "Pawnee", "in", 3175553046, "perd@journalpawnee.example.com"),
]

# (src_type, src_name, rel_type, dst_type, dst_name, title)
PAWNEE_RELATIONSHIPS = [
    ("Human", "Thomas Haverford", "Owner of", "Entity", "Tom’s Bistro", None),
    ("Entity", "Tom’s Bistro", "Owned by", "Human", "Thomas Haverford", None),
    ("Entity", "Sweetums Candy Factory", "Subsidiary of", "Entity", "Sweetums", None),
    ("Human", "Kyle Nichols", "Employee of", "Entity", "Sweetums Candy Factory", "Plant Worker"),
    ("Entity", "Sweetums Candy Factory", "Employer of", "Human", "Kyle Nichols", "Plant Worker"),
    ("Human", "Evelyn Stonefield", "Owner of", "Entity", "Sweetums Candy Factory", None),
    ("Human", "Teddy Blaszczak", "Employee of", "Entity", "Pawnee Commons Construction", "Foreman"),
    ("Entity", "Pawnee Commons Construction", "Employer of", "Human", "Teddy Blaszczak", "Foreman"),
    ("Human", "Ann Perkins", "Owner of", "Entity", "Ann Perkins Wellness Clinic", None),
    ("Entity", "Ann Perkins Wellness Clinic", "Owned by", "Human", "Ann Perkins", None),
    ("Human", "Craig Middlebrooks", "Employee of", "Entity", "Pawnee Parks & Recreation Department", "Assistant Director"),
    ("Entity", "Pawnee Parks & Recreation Department", "Employer of", "Human", "Craig Middlebrooks", "Assistant Director"),
    ("Human", "Tynnyfer McGraw", "Employee of", "Entity", "Pawnee Parks & Recreation Department", "Assistant"),
    ("Human", "Diane Lewis", "Employee of", "Entity", "Eagleton Country Club", "Librarian"),
    ("Human", "Tammy Swanson", "Employee of", "Entity", "Eagleton Country Club", "Head Librarian"),
    ("Human", "Burt Macklin", "Employee of", "Entity", "Pawnee Parks & Recreation Department", "Field Agent"),
    ("Human", "Jennifer Barkley-Wyatt", "Spouse of", "Human", "Benjamin Wyatt", None),
    ("Human", "Andy Dwyer", "Co-Worker of", "Human", "April Ludgate", None),
    ("Human", "Perd Hapley", "Employee of", "Entity", "The Pawnee Journal", "Anchor"),
    ("Entity", "The Pawnee Journal", "Employer of", "Human", "Perd Hapley", "Anchor"),
]

# ---------- ANNAPOLIS / CHESAPEAKE UNIVERSE (Coastal_Harbor_Export.xlsx) ---

COASTAL_ENTITIES = [
    ("ent018", "Tidewater Coffee Roasters", "llc", dt.datetime(2012, 4, 1),
     "Specialty coffee roaster with retail café on Main Street", 1_600_000, False,
     "Tidewater Coffee", "52-3344556", None, True),
    ("ent019", "Chesapeake Machine Works", "c_Corp", dt.datetime(1967, 3, 15),
     "Precision CNC machining for marine and aerospace parts", 27_500_000, False,
     None, "52-1122334", None, True),
    ("ent020", "Harbor Build & Trade", "llc", dt.datetime(2008, 11, 12),
     "General contractor specializing in dockside and waterfront renovations", 5_100_000, False,
     "Harbor Build", "52-2233445", None, True),
    ("ent021", "Annapolis Direct Primary Care", "sole_Proprietor", dt.datetime(2017, 6, 20),
     "Membership-based primary care clinic", 520_000, False,
     None, None, "412-55-7788", True),
    ("ent022", "Bayside Cloud Services", "c_Corp", dt.datetime(2015, 9, 1),
     "Regional B2B cloud-hosting startup; SOC2-Type-2 certified", 12_400_000, False,
     "Bayside Cloud", "27-9988112", None, True),
    ("ent023", "Severn River Yacht Club", "limited_Partnership", dt.datetime(1956, 5, 1),
     "Private yacht club and sailing school on the Severn River", 7_800_000, False,
     None, "52-4455667", None, True),
    ("ent024", "Chesapeake Bay Aquarium", "not_For_Profit", dt.datetime(1988, 7, 1),
     "Regional aquarium and marine education nonprofit", 1_100_000, False,
     None, "52-6677889", None, True),
    ("ent025", "Eastport Crab Company", "llc", dt.datetime(2001, 3, 5),
     "Seafood wholesaler and steamed-crab retail market", 2_300_000, False,
     "Eastport Crabs", "52-5566778", None, True),
    ("ent026", "Naptown Brewing", "c_Corp", dt.datetime(2014, 2, 14),
     "Craft brewery and taproom in downtown Annapolis", 3_600_000, False,
     "Naptown Brewing", "52-7788990", None, True),
    ("ent027", "Harbor Catering", "sole_Proprietor", dt.datetime(2019, 8, 10),
     "Event catering spin-off from Coastal Grille & Burgers", 210_000, False,
     "Coastal Catering", None, "412-55-7790", True),
]

COASTAL_HUMANS = [
    ("hum032", "Mr.", "Theodore", None, "Blount", "Ted", "hehim",
     dt.datetime(1961, 9, 12), "high_School", "skilled_Trade_Technician",
     "construction", 1980, False, "male", "married",
     "412-56-7891", None, None, None, True, None, None, None, None, None),
    ("hum033", "Ms.", "Monica", None, "Serrano", "Monica", "sheher",
     dt.datetime(1987, 6, 3), "bachelors", "sales_Marketing",
     "hospitality", 2009, False, "female", "single",
     "412-67-8912", None, None, None, True, None, None, None, None, None),
    ("hum034", "Dr.", "Richard", None, "Okafor", "Rick", "hehim",
     dt.datetime(1959, 11, 28), "phd", "healthcare_Professional",
     "healthcare", 1987, False, "male", "married",
     "412-78-9123", None, None, None, True, None, None, None, None, None),
    ("hum035", "Mr.", "Craig", None, "Maddox", "Craig", "hehim",
     dt.datetime(1973, 2, 17), "bachelors", "manager_Supervisor",
     "government", 2003, False, "male", "single",
     "412-89-1234", None, None, None, True, None, None, None, None, None),
    ("hum036", "Ms.", "Tiffany", None, "Marsh", "Tiff", "sheher",
     dt.datetime(1990, 4, 22), "bachelors", "administrator_Clerical",
     "government", 2015, False, "female", "single",
     "412-91-2345", None, None, None, True, None, None, None, None, None),
    ("hum037", "Mr.", "Kyle", None, "Hendricks", "Kyle", "hehim",
     dt.datetime(1981, 10, 10), "high_School", "laborer",
     "manufacturing", 2002, False, "male", "single",
     "412-12-3456", None, None, None, True, None, None, None, None, None),
    ("hum038", "Ms.", "Brandi", None, "Ortiz", "Brandi", "sheher",
     dt.datetime(1984, 7, 8), "high_School", "arts_Entertainment_Recreation",
     "entertainment", 2004, False, "female", "single",
     "412-23-4567", None, None, None, True, None, None, None, None, None),
    ("hum039", "Mr.", "Oren", None, None, None, "hehim",
     dt.datetime(1976, 12, 1), "masters", "other",
     "arts", 2006, False, "male", "single",
     "412-34-5678", None, None, None, False, None, None, None, None, None),
    ("hum040", "Mrs.", "Tamara", "J.", "Sutton", "Tam", "sheher",
     dt.datetime(1969, 8, 24), "masters", "administrator_Clerical",
     "education", 1996, False, "female", "divorced",
     "412-45-6789", None, None, None, True, None, None, None, None, None),
    ("hum041", "Ms.", "Diana", None, "Leighton", "Diana", "sheher",
     dt.datetime(1971, 3, 11), "masters", "educator_Professor",
     "education", 2001, False, "female", "married",
     "412-56-7892", None, None, None, True, None, None, None, None, None),
    ("hum042", "Mr.", "Burton", None, "Macklin", "Burt", "hehim",
     dt.datetime(1978, 5, 14), "bachelors", "legal_Financial_Professional",
     "government", 2005, False, "male", "married",
     "412-67-8913", None, None, None, True, None, None, None, None, None),
    ("hum043", "Ms.", "Evelyn", None, "Stonefield", "Evelyn", "sheher",
     dt.datetime(1964, 1, 29), "bachelors", "executive_Owner",
     "manufacturing", 1989, False, "female", "married",
     "412-78-9124", None, None, None, True, None, None, None, None, None),
    ("hum044", "Mr.", "Andrew", None, "Pryor", "Drew", "hehim",
     dt.datetime(1983, 4, 6), "high_School", "arts_Entertainment_Recreation",
     "entertainment", 2004, False, "male", "married",
     "412-89-1235", None, None, None, True, None, None, None, None, None),
    ("hum045", "Mrs.", "Jennifer", None, "Barkley-Holt", "Jen", "sheher",
     dt.datetime(1975, 2, 23), "masters", "sales_Marketing",
     "politics", 1999, False, "female", "single",
     "412-91-2346", None, None, None, True, None, None, None, None, None),
    ("hum046", "Mr.", "Perry", None, "Harper", "Perry", "hehim",
     dt.datetime(1965, 10, 17), "bachelors", "information_Media_Communications",
     "media", 1989, False, "male", "married",
     "412-12-3457", None, None, None, True, None, None, None, None, None),
]

_COASTAL_CONTACTS_ENTITY = [
    ("ent018", "Café", "140 Main Street", "Annapolis", "md", 4105552018, "https://tidewatercoffee.example.com", "hello@tidewatercoffee.example.com"),
    ("ent019", "Plant", "2200 Industrial Park Rd", "Annapolis", "md", 4105552019, "https://chesapeakemachineworks.example.com", "orders@cmw.example.com"),
    ("ent020", "Yard", "88 Dockside Ln", "Annapolis", "md", 4105552020, "https://harborbuild.example.com", "jobs@harborbuild.example.com"),
    ("ent021", "Clinic", "500 West Street", "Annapolis", "md", 4105552021, "https://annapolisdpc.example.com", "info@annapolisdpc.example.com"),
    ("ent022", "HQ", "900 Bay Ridge Ave", "Annapolis", "md", 4105552022, "https://baysidecloud.example.com", "sales@baysidecloud.example.com"),
    ("ent023", "Clubhouse", "1 Severn Point Rd", "Annapolis", "md", 4105552023, "https://severnriveryc.example.com", "events@severnriveryc.example.com"),
    ("ent024", "Main Entry", "50 Aquarium Way", "Annapolis", "md", 4105552024, "https://chesapeakebayaquarium.example.com", "info@chesapeakebayaquarium.example.com"),
    ("ent025", "Market", "11 Dock Street", "Annapolis", "md", 4105552025, "https://eastportcrab.example.com", "orders@eastportcrab.example.com"),
    ("ent026", "Taproom", "77 West Street", "Annapolis", "md", 4105552026, "https://naptownbrewing.example.com", "hello@naptownbrewing.example.com"),
    ("ent027", "HQ", "120 Harbor Dr", "Annapolis", "md", 4105552027, None, "catering@coastalgrille.example.com"),
]

_COASTAL_CONTACTS_HUMAN_PERSONAL = [
    ("hum032", "15 Lumber Ln", "Annapolis", "md", 4105553032, "ted.blount@example.com"),
    ("hum033", "220 Dock St", "Annapolis", "md", 4105553033, "monica.serrano@example.com"),
    ("hum034", "500 West Street", "Annapolis", "md", 4105553034, "dr.okafor@example.com"),
    ("hum035", "140 Harbor Dr", "Annapolis", "md", 4105553035, "craig.maddox@annapolisparks.example.com"),
    ("hum036", "140 Harbor Dr", "Annapolis", "md", 4105553036, "tiff@annapolis.example.com"),
    ("hum037", "2200 Industrial Park Rd", "Annapolis", "md", 4105553037, "kyle.h@cmw.example.com"),
    ("hum038", "220 Dock St", "Annapolis", "md", 4105553038, "brandi.ortiz@example.com"),
    ("hum039", "77 Artisan Way", "Annapolis", "md", 4105553039, "oren@example.com"),
    ("hum040", "900 Library Row", "Annapolis", "md", 4105553040, "tamara@annapolislib.example.com"),
    ("hum041", "1 College Ave", "Annapolis", "md", 4105553041, "diana.leighton@example.com"),
    ("hum042", "140 Harbor Dr", "Annapolis", "md", 4105553042, "burt.macklin@annapolis.example.com"),
    ("hum043", "2200 Industrial Park Rd", "Annapolis", "md", 4105553043, "evelyn@cmw.example.com"),
    ("hum044", "18 Waterman Way", "Annapolis", "md", 4105553044, "drew.pryor@example.com"),
    ("hum045", "450 Campaign Ln", "Annapolis", "md", 4105553045, "jen.holt@example.com"),
    ("hum046", "100 Press Row", "Annapolis", "md", 4105553046, "perry@annapolissentinel.example.com"),
]

COASTAL_RELATIONSHIPS = [
    ("Human", "Monica Serrano", "Owner of", "Entity", "Tidewater Coffee Roasters", None),
    ("Entity", "Tidewater Coffee Roasters", "Owned by", "Human", "Monica Serrano", None),
    ("Entity", "Chesapeake Machine Works", "Subsidiary of", "Entity", "Stonefield Confections", None),
    ("Human", "Kyle Hendricks", "Employee of", "Entity", "Chesapeake Machine Works", "Machinist"),
    ("Entity", "Chesapeake Machine Works", "Employer of", "Human", "Kyle Hendricks", "Machinist"),
    ("Human", "Evelyn Stonefield", "Owner of", "Entity", "Chesapeake Machine Works", None),
    ("Human", "Theodore Blount", "Employee of", "Entity", "Harbor Build & Trade", "Foreman"),
    ("Entity", "Harbor Build & Trade", "Employer of", "Human", "Theodore Blount", "Foreman"),
    ("Human", "Richard Okafor", "Owner of", "Entity", "Annapolis Direct Primary Care", None),
    ("Entity", "Annapolis Direct Primary Care", "Owned by", "Human", "Richard Okafor", None),
    ("Human", "Craig Maddox", "Employee of", "Entity", "Annapolis Parks Department", "Assistant Director"),
    ("Entity", "Annapolis Parks Department", "Employer of", "Human", "Craig Maddox", "Assistant Director"),
    ("Human", "Tiffany Marsh", "Employee of", "Entity", "Annapolis Parks Department", "Assistant"),
    ("Human", "Diana Leighton", "Employee of", "Entity", "Severn River Yacht Club", "Sailing Instructor"),
    ("Human", "Tamara Sutton", "Employee of", "Entity", "Chesapeake Historical Society", "Director"),
    ("Human", "Burton Macklin", "Employee of", "Entity", "Annapolis Parks Department", "Investigator"),
    ("Human", "Andrew Pryor", "Co-Worker of", "Human", "Drew Pryor", None),
    ("Human", "Perry Harper", "Employee of", "Entity", "The Annapolis Sentinel", "Anchor"),
    ("Entity", "The Annapolis Sentinel", "Employer of", "Human", "Perry Harper", "Anchor"),
    ("Human", "Jennifer Barkley-Holt", "Consultant to", "Entity", "Bayside Cloud Services", None),
]

# --------------------- builders ------------------------------------------

def _entity_contact_row(ent_id, name_hint, line1, city, state, phone, website, email):
    """Build a 40-col contact row for an entity (BUSINESS contactType)."""
    return (
        "entity", ent_id, "BUSINESS", None, None, None, name_hint, "email", False,
        line1, None, None, None, city, state, None, "us",
        dt.datetime(2018, 1, 1),
        line1, None, None, None, city, state, None, "us",
        phone, 1, True, None,     # primary phone
        None, None, None, None,   # secondary phone
        None, None,               # fax
        website, None, email, None,
    )


def _human_personal_contact_row(hum_id, line1, city, state, phone, email):
    """Build a 40-col contact row for a human (PERSONAL contactType)."""
    return (
        "human", hum_id, "PERSONAL", None, None, None, None, "email", False,
        line1, None, None, None, city, state, None, "us",
        dt.datetime(2018, 1, 1),
        line1, None, None, None, city, state, None, "us",
        phone, 1, True, None,
        None, None, None, None,
        None, None,
        None, None, email, None,
    )


def _extend_workbook(path: Path, entities, humans, ce, ch, rels) -> dict:
    wb = openpyxl.load_workbook(path)
    report = {"file": path.name}

    ws_e = wb["Entity Demo Data"]
    report["entities_added"] = _append(ws_e, entities, id_col=1)

    ws_h = wb["Human Demo Data"]
    report["humans_added"] = _append(ws_h, humans, id_col=1)

    ws_c = wb["Contact Demo Data"]
    ent_contact_rows = [_entity_contact_row(*args) for args in ce]
    hum_contact_rows = [_human_personal_contact_row(*args) for args in ch]
    # Contacts have no stable PK; append unconditionally but guarded by (type, id) existence.
    existing_contact_keys = {
        (r[0], r[1], r[2], r[9]) for r in ws_c.iter_rows(min_row=2, values_only=True) if r[1]
    }
    added = 0
    for row in ent_contact_rows + hum_contact_rows:
        key = (row[0], row[1], row[2], row[9])
        if key in existing_contact_keys:
            continue
        ws_c.append(list(row))
        added += 1
    report["contacts_added"] = added

    ws_r = wb["Relationship Demo Data"]
    existing_rels = {
        tuple((c or "").strip() for c in row[:5])
        for row in ws_r.iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    added = 0
    for r in rels:
        key = tuple((c or "").strip() for c in r[:5])
        if key in existing_rels:
            continue
        ws_r.append(list(r))
        added += 1
    report["relationships_added"] = added

    wb.save(path)
    return report


def main():
    pawnee = _extend_workbook(
        DATA / "demo.xlsx",
        PAWNEE_ENTITIES, PAWNEE_HUMANS,
        _PAWNEE_CONTACTS_ENTITY, _PAWNEE_CONTACTS_HUMAN_PERSONAL,
        PAWNEE_RELATIONSHIPS,
    )
    coastal = _extend_workbook(
        DATA / "Coastal_Harbor_Export.xlsx",
        COASTAL_ENTITIES, COASTAL_HUMANS,
        _COASTAL_CONTACTS_ENTITY, _COASTAL_CONTACTS_HUMAN_PERSONAL,
        COASTAL_RELATIONSHIPS,
    )
    for rep in (pawnee, coastal):
        print(f"[{rep['file']}] +{rep['entities_added']} entities, "
              f"+{rep['humans_added']} humans, "
              f"+{rep['contacts_added']} contacts, "
              f"+{rep['relationships_added']} relationships")


if __name__ == "__main__":
    main()
