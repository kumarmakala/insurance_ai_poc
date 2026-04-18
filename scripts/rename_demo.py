"""Rewrite data/demo.xlsx in-place to replace TV-character data with a
plausible insurance-agency book for a CEO demo. Preserves identifiers,
addresses, and sheet structure so the rest of the pipeline still works."""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "demo.xlsx"

# --- Master rename maps --------------------------------------------------

ENTITY_NAMES = {
    "Bob's Burgers": "Coastal Grille & Burgers",
    "Bob\u2019s Burgers": "Coastal Grille & Burgers",
    "Jimmy Pesto's Pizzeria": "Romano's Brick Oven",
    "Jimmy Pesto\u2019s Pizzeria": "Romano's Brick Oven",
    "Wagstaff School": "Hillcrest Academy",
    "Wonder Wharf": "Harbor Pier Amusements",
    "Sweetums": "Stonefield Confections",
    "Pawnee Parks & Recreation Department": "Annapolis Parks Department",
    "Pawnee Parks & Recreation": "Annapolis Parks Department",
    "Pawnee City Hall": "City of Annapolis",
    "City of Pawnee": "City of Annapolis",
    "JJ's Diner": "Main Street Diner",
    "JJ\u2019s Diner": "Main Street Diner",
    "Entertainment 720": "Vistria Media Group",
    "Rent-A-Swag": "Midtown Wardrobe Rentals",
    "The Pawnee Journal": "The Annapolis Sentinel",
    "Paunch Burger": "Hometown Burgers",
    "The Snakehole Lounge": "The Rook Tavern",
    "Ron Swanson Fine Woodworking": "Donovan Fine Woodworking",
    "Pawnee Historical Society": "Chesapeake Historical Society",
    "Donna Meagle's Real Estate Services": "Morgan Real Estate Services",
    "Donna Meagle\u2019s Real Estate Services": "Morgan Real Estate Services",
    "Pawnee St. Joseph's Hospital": "St. Joseph's Annapolis Hospital",
    "Pawnee St. Joseph\u2019s Hospital": "St. Joseph's Annapolis Hospital",
    "Pawnee St. Josephs Hospital": "St. Joseph's Annapolis Hospital",
    "Tilton & Radomski Accounting": "Preston & Hartwell Accounting",
    "Tilton & Radomski Accountiing": "Preston & Hartwell Accounting",
    "Very Good Building & Development": "Blackfin Building & Development",
}

# (first, middle, last) → (new_first, new_middle, new_last, new_preferred)
HUMAN_MAP = {
    ("Bob", None, "Belcher"): ("Robert", None, "Chen", "Rob"),
    ("Linda", None, "Belcher"): ("Linda", None, "Chen", "Linda"),
    ("Tina", "Ruth", "Belcher"): ("Tyler", "Ruth", "Chen", "Ty"),
    ("Gene", "Tobias", "Belcher"): ("Grant", "Tobias", "Chen", "Grant"),
    ("Louise", None, "Belcher"): ("Leah", None, "Chen", "Leah"),
    ("Jimmy", None, "Pesto Jr."): ("James", None, "Romano Jr.", "Jim"),
    ("Leslie", "Barbara", "Knope"): ("Laura", "Barbara", "Kline", "Laura"),
    ("Ronald", "Ulysses", "Swanson"): ("Ronald", "Ulysses", "Donovan", "Ron"),
    ("Thomas", "Montgomery", "Haverford"): ("Thomas", "Montgomery", "Harlow", "Tom"),
    ("April", "Roberta", "Ludgate"): ("April", "Roberta", "Lindgren", "April"),
    ("Benjamin", "Scott", "Wyatt"): ("Benjamin", "Scott", "Ward", "Ben"),
    ("Ann", "Meredith", "Perkins"): ("Ann", "Meredith", "Parker", "Ann"),
    ("Andrew", "Burly", "Dwyer"): ("Andrew", "Burke", "Doyle", "Andy"),
    ("Christopher", "Mark", "Traeger"): ("Christopher", "Mark", "Traynor", "Chris"),
    ("Donna", "Marie", "Meagle"): ("Danielle", "Marie", "Morgan", "Dani"),
    ("Jerry", "Gengurch", "Gergich"): ("Jerome", "Gerard", "Gallagher", "Jerry"),
    ("Perd", None, "Hapley"): ("Paul", None, "Hadley", "Paul"),
    ("Joan", None, "Callamezzo"): ("Joan", None, "Calloway", "Joan"),
    ("Shauna", None, "Malwae-Tweep"): ("Shauna", None, "Malloy-Tripp", "Shauna"),
    ("Jennifer", None, "Barkley"): ("Jennifer", None, "Barker", "Jen"),
    ("Jean-Ralphio", None, "Saperstein"): ("Jonathan", None, "Sapirstein", "Jon"),
    ("Mona-Lisa", None, "Saperstein"): ("Monica", None, "Sapirstein", "Monica"),
    ("Ethel", None, "Beavers"): ("Ethel", None, "Bowers", "Ethel"),
    ("Harris", None, "Humphrey"): ("Harrison", None, "Humboldt", "Harry"),
    ("Donna", None, "St. James"): ("Deborah", None, "Stedman", "Deb"),
    ("Ingrid", None, "de Forest"): ("Ingrid", None, "Dereksson", "Ingrid"),
    ("Jeremy", None, "Jamm"): ("Jeremy", None, "Jameson", "Jeremy"),
    ("Trevor", None, "Nelsson"): ("Trevor", None, "Nielsen", "Trevor"),
    ("Kyle", None, "Nichols"): ("Kyle", None, "Nichols", "Kyle"),
    ("Marcia", "Langman", "Langman"): ("Marcia", None, "Langley", "Marcia"),
    ("Councilman", "Howser", "Jammerson"): ("Howard", None, "Jameson", "Howard"),
}

# Full-name strings that appear in Relationship Demo Data (Human type)
HUMAN_FULL_NAMES = {
    "Bob Belcher": "Robert Chen",
    "Linda Belcher": "Linda Chen",
    "Tina Belcher": "Tyler Chen",
    "Gene Belcher": "Grant Chen",
    "Louise Belcher": "Leah Chen",
    "Jimmy Pesto Jr.": "James Romano Jr.",
    "Leslie Knope": "Laura Kline",
    "Ronald Swanson": "Ronald Donovan",
    "Ron Swanson": "Ronald Donovan",
    "Thomas Haverford": "Thomas Harlow",
    "Tom Haverford": "Thomas Harlow",
    "April Ludgate": "April Lindgren",
    "Benjamin Wyatt": "Benjamin Ward",
    "Ben Wyatt": "Benjamin Ward",
    "Ann Perkins": "Ann Parker",
    "Anne Perkins": "Ann Parker",
    "Andrew Dwyer": "Andrew Doyle",
    "Andy Dwyer": "Andrew Doyle",
    "Christopher Traeger": "Christopher Traynor",
    "Chris Traeger": "Christopher Traynor",
    "Donna Meagle": "Danielle Morgan",
    "Jerry Gergich": "Jerome Gallagher",
    "Perd Hapley": "Paul Hadley",
    "Joan Callamezzo": "Joan Calloway",
    "Shauna Malwae-Tweep": "Shauna Malloy-Tripp",
    "Jennifer Barkley": "Jennifer Barker",
    "Jen Barkley": "Jennifer Barker",
    "Jean-Ralphio Saperstein": "Jonathan Sapirstein",
    "Mona-Lisa Saperstein": "Monica Sapirstein",
    "Ethel Beavers": "Ethel Bowers",
    "Harris Humphrey": "Harrison Humboldt",
    "Donna St. James": "Deborah Stedman",
    "Ingrid de Forest": "Ingrid Dereksson",
    "Jeremy Jamm": "Jeremy Jameson",
    "Trevor Nelsson": "Trevor Nielsen",
    "Marcia Langman": "Marcia Langley",
    # Phantom refs only in relationships (not in Human sheet)
    "Marlene Griggs Knope": "Marlene Griggs-Kline",
    "Marlene Griggs-Knope": "Marlene Griggs-Kline",
    "Diane Lewis": "Diane Lewis",
    "Joan Johnson (JJ)": "Joan Johnson",
    "Joan Johnson": "Joan Johnson",
}

CITY_MAP = {
    "Pawnee": "Annapolis",
    "Seymour's Bay": "Annapolis",
    "Seymour\u2019s Bay": "Annapolis",
    "Seymours Bay": "Annapolis",
}

STATE_MAP = {
    "in": "md",  # Pawnee, IN → Annapolis, MD
    "ca": "md",  # Seymour's Bay, CA → Annapolis, MD
}

# Description free-text substitutions (applied in order, regex-safe)
TEXT_SUBS = [
    ("Seymour\u2019s Bay", "Annapolis"),
    ("Seymour's Bay", "Annapolis"),
    ("Seymours Bay", "Annapolis"),
    ("Pawnee Journal", "Annapolis Sentinel"),
    ("Pawnee", "Annapolis"),
    ("Bob\u2019s Burgers", "Coastal Grille & Burgers"),
    ("Bob's Burgers", "Coastal Grille & Burgers"),
    ("Jimmy Pesto\u2019s", "Romano\u2019s"),
    ("Jimmy Pesto's", "Romano's"),
    ("Jimmy Pesto", "Romano's"),
    ("Snakehole Lounge", "Rook Tavern"),
    ("Swanson Fine Woodworking", "Donovan Fine Woodworking"),
    ("Meagle Real Estate", "Morgan Real Estate"),
    ("Tom Haverford", "Thomas Harlow"),
    ("Jean-Ralphio Saperstein", "Jonathan Sapirstein"),
    ("Jean-Ralphio", "Jonathan"),
    ("Donna Meagle", "Danielle Morgan"),
    ("Sweetums", "Stonefield Confections"),
    ("Leslie Knope", "Laura Kline"),
    ("Ron Swanson", "Ronald Donovan"),
]


def apply_text_subs(val):
    if not isinstance(val, str):
        return val
    for old, new in TEXT_SUBS:
        val = val.replace(old, new)
    return val


def rewrite_entities(sheet):
    dba_idx = None
    for idx, cell in enumerate(sheet[1]):
        if cell.value == "doingBusinessAs":
            dba_idx = idx
    for row in sheet.iter_rows(min_row=2):
        if not row[0].value:
            continue
        # column 1 = name
        name = row[1].value
        if name in ENTITY_NAMES:
            row[1].value = ENTITY_NAMES[name]
        # dba — try exact map, then fall back to text sub
        if dba_idx is not None:
            dba = row[dba_idx].value
            if dba in ENTITY_NAMES:
                row[dba_idx].value = ENTITY_NAMES[dba]
            else:
                row[dba_idx].value = apply_text_subs(dba)
        # entityDescription (col 4) — free text replacement
        if len(row) > 4:
            row[4].value = apply_text_subs(row[4].value)


def rewrite_humans(sheet):
    headers = [c.value for c in sheet[1]]
    fi = headers.index("firstName")
    mi = headers.index("middleName")
    li = headers.index("lastName")
    pi = headers.index("preferredName")
    for row in sheet.iter_rows(min_row=2):
        if not row[0].value:
            continue
        first = row[fi].value
        mid = row[mi].value
        last = row[li].value
        key = (first, mid, last)
        if key in HUMAN_MAP:
            nf, nm, nl, np = HUMAN_MAP[key]
            row[fi].value = nf
            row[mi].value = nm
            row[li].value = nl
            row[pi].value = np
        else:
            # fall back to partial match by (first, last) ignoring middle
            fallback = next(((k, v) for k, v in HUMAN_MAP.items() if k[0] == first and k[2] == last), None)
            if fallback:
                _, (nf, nm, nl, np) = fallback
                row[fi].value = nf
                row[mi].value = nm if nm else row[mi].value
                row[li].value = nl
                row[pi].value = np


def rewrite_contacts(sheet):
    headers = [c.value for c in sheet[1]]
    name_idx = headers.index("name") if "name" in headers else None
    city_idx = headers.index("physicalAddress_city") if "physicalAddress_city" in headers else None
    state_idx = headers.index("physicalAddress_state") if "physicalAddress_state" in headers else None
    mail_city_candidates = [i for i, h in enumerate(headers) if h and "mailing" in h.lower() and "city" in h.lower()]
    mail_state_candidates = [i for i, h in enumerate(headers) if h and "mailing" in h.lower() and "state" in h.lower()]
    for row in sheet.iter_rows(min_row=2):
        if not row[0].value:
            continue
        if name_idx is not None:
            row[name_idx].value = apply_text_subs(row[name_idx].value)
        if city_idx is not None:
            c = row[city_idx].value
            if c in CITY_MAP:
                row[city_idx].value = CITY_MAP[c]
        if state_idx is not None:
            s = row[state_idx].value
            if isinstance(s, str) and s.lower() in STATE_MAP:
                row[state_idx].value = STATE_MAP[s.lower()]
        for idx in mail_city_candidates:
            v = row[idx].value
            if v in CITY_MAP:
                row[idx].value = CITY_MAP[v]
        for idx in mail_state_candidates:
            v = row[idx].value
            if isinstance(v, str) and v.lower() in STATE_MAP:
                row[idx].value = STATE_MAP[v.lower()]


def rewrite_relationships(sheet):
    # headers: Record Type, Name, Relationship, Record Type, Name 2, Title
    for row in sheet.iter_rows(min_row=2):
        rt1 = row[0].value
        n1 = row[1].value
        rt2 = row[3].value
        n2 = row[4].value
        if rt1 == "Human" and n1 in HUMAN_FULL_NAMES:
            row[1].value = HUMAN_FULL_NAMES[n1]
        elif rt1 == "Entity" and n1 in ENTITY_NAMES:
            row[1].value = ENTITY_NAMES[n1]
        if rt2 == "Human" and n2 in HUMAN_FULL_NAMES:
            row[4].value = HUMAN_FULL_NAMES[n2]
        elif rt2 == "Entity" and n2 in ENTITY_NAMES:
            row[4].value = ENTITY_NAMES[n2]


def main():
    wb = openpyxl.load_workbook(XLSX)
    rewrite_entities(wb["Entity Demo Data"])
    rewrite_humans(wb["Human Demo Data"])
    rewrite_contacts(wb["Contact Demo Data"])
    rewrite_relationships(wb["Relationship Demo Data"])
    wb.save(XLSX)
    print(f"Rewrote {XLSX}")


if __name__ == "__main__":
    main()
